from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

REQUIRED_COLUMNS = ["timestamp", "service", "level", "message", "response_ms"]

# -------------------------
# Load + validate
# -------------------------


def load_logs(path: Path) -> pd.DataFrame:
    """
    Load a CSV file into a normalized DataFrame.

    - Validates schema
    - Parses timestamp and response_ms
    - Normalizes `level` to uppercase
    - Sorts by timestamp for stable reporting
    """
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    df = pd.read_csv(path)
    validate_logs(df)

    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df["response_ms"] = pd.to_numeric(df["response_ms"], errors="coerce")

    # Normalize for consistent reporting and filtering
    df["level"] = df["level"].astype(str).str.upper()
    df["service"] = df["service"].astype(str)

    df = df.sort_values("timestamp", kind="mergesort").reset_index(drop=True)
    return df


def validate_logs(df: pd.DataFrame) -> None:
    """
    Validate required CSV schema.

    Raises:
        ValueError: if any required columns are missing
    """
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            "CSV schema invalid. Missing columns: "
            + ", ".join(missing)
            + f"\nExpected columns: {', '.join(REQUIRED_COLUMNS)}"
        )


# -------------------------
# Filtering
# -------------------------


def apply_filters(df: pd.DataFrame, service: str | None, level: str | None) -> pd.DataFrame:
    """
    Apply optional filters (service, level).

    Note:
        `level` matching is case-insensitive (we normalize to upper).
    """
    filtered = df

    if service:
        filtered = filtered[filtered["service"] == str(service)]

    if level:
        filtered = filtered[filtered["level"] == str(level).upper()]

    return filtered.reset_index(drop=True)


# -------------------------
# Report dataframes
# -------------------------


def build_export_logs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the logs export DataFrame for Excel.

    Output columns:
        date, time, service, level, message, response_ms
    """
    out = df.copy()
    out["date"] = out["timestamp"].dt.date
    out["time"] = out["timestamp"].dt.strftime("%H:%M:%S")
    return out[["date", "time", "service", "level", "message", "response_ms"]]


def build_summary_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Build the summary tables:
        - Key metrics
        - Counts by level
        - Counts by service
    """
    total_rows = len(df)
    error_count = int((df["level"] == "ERROR").sum())

    summary_df = pd.DataFrame(
        [
            {"metric": "total_rows", "value": total_rows},
            {"metric": "error_count", "value": error_count},
        ]
    )

    per_level = (
        df["level"].value_counts(dropna=False).rename_axis("level").reset_index(name="count")
    )
    per_service = (
        df["service"].value_counts(dropna=False).rename_axis("service").reset_index(name="count")
    )

    return summary_df, per_level, per_service


def build_daily_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a daily pivot summary with counts by level and derived totals.

    Output columns:
        date, total_rows, error_count, <LEVEL columns...>
    """
    daily_pivot = (
        df.assign(date=df["timestamp"].dt.date)
        .pivot_table(index="date", columns="level", values="message", aggfunc="count", fill_value=0)
        .reset_index()
        .sort_values("date")
    )

    level_cols = [c for c in daily_pivot.columns if c != "date"]
    daily_pivot["total_rows"] = daily_pivot[level_cols].sum(axis=1) if level_cols else 0
    daily_pivot["error_count"] = daily_pivot["ERROR"] if "ERROR" in daily_pivot.columns else 0

    ordered_cols = ["date", "total_rows", "error_count"] + [
        c for c in level_cols if c not in ("total_rows", "error_count")
    ]
    return daily_pivot[ordered_cols]


# -------------------------
# Excel styling (post-save)
# -------------------------


def _autofit_columns(ws, max_width: int = 60, scan_limit: int = 2000) -> None:
    """
    Set column widths based on content length (best-effort).

    Args:
        ws: openpyxl worksheet
        max_width: clamp column width to this maximum
        scan_limit: max number of cells scanned per column (performance guard)
    """
    for col in ws.columns:
        first_cell = col[0]
        if not hasattr(first_cell, "column_letter"):
            continue

        col_letter = first_cell.column_letter
        max_len = 0

        for cell in col[:scan_limit]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)

        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _style_header_row(ws, header_row: int = 1) -> None:
    """
    Apply a clean 'table header' style to a specific worksheet row.
    """
    fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    ws.row_dimensions[header_row].height = 18


def _add_section_title(ws, row: int, title: str) -> None:
    """
    Add a bold section title (single cell) at a given row.
    """
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    ws.row_dimensions[row].height = 18


def _highlight_error_rows(
    ws,
    level_col_name: str = "level",
    header_row: int = 1,
) -> None:
    """
    Highlight entire rows where the `level` column equals 'ERROR'.

    Uses a relative row reference so Excel expands it correctly.
    """

    # Extract header names
    headers = [cell.value for cell in ws[header_row]]
    if level_col_name not in headers:
        return

    level_col_index = headers.index(level_col_name) + 1
    level_col_letter = ws.cell(row=header_row, column=level_col_index).column_letter

    start_row = header_row + 1
    end_row = ws.max_row

    if end_row < start_row:
        return

    last_col_letter = ws.cell(row=header_row, column=ws.max_column).column_letter
    data_range = f"A{start_row}:{last_col_letter}{end_row}"

    fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # IMPORTANT:
    # Use row reference WITHOUT fixing row number
    # This makes it behave like Excel's = $D2="ERROR"
    formula = f'${level_col_letter}{start_row}="ERROR"'

    rule = FormulaRule(formula=[formula], fill=fill)

    ws.conditional_formatting.add(data_range, rule)


def style_workbook(xlsx_path: Path, per_level_len: int) -> None:
    """
    Apply workbook styling after the XLSX has been written.

    Styling rules:
      - logs: freeze header, filter, header style, highlight ERROR rows, autofit
      - daily_summary: freeze header, filter, header style, autofit
      - summary: add section titles, style EACH table header row, autofit
    """
    wb = load_workbook(xlsx_path)

    # logs
    if "logs" in wb.sheetnames:
        ws = wb["logs"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        _style_header_row(ws, 1)
        _highlight_error_rows(ws, "level", 1)
        _autofit_columns(ws)

    # daily_summary
    if "daily_summary" in wb.sheetnames:
        ws = wb["daily_summary"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        _style_header_row(ws, 1)
        _autofit_columns(ws)

    # summary (multi-table sheet)
    if "summary" in wb.sheetnames:
        ws = wb["summary"]

        # Layout (Excel rows):
        # 1: title "Key Metrics"
        # 2: summary_df header
        # 3.. : summary_df data (2 rows)
        # 6: title "Counts by Level"
        # 7: per_level header
        # 8.. : per_level data (per_level_len rows)
        # (blank row)
        # 9 + per_level_len: title "Counts by Service"
        # 10 + per_level_len: per_service header
        key_metrics_title_row = 1
        key_metrics_header_row = 2

        counts_by_level_title_row = 6
        counts_by_level_header_row = 7

        counts_by_service_title_row = 9 + per_level_len
        counts_by_service_header_row = 10 + per_level_len

        _add_section_title(ws, row=key_metrics_title_row, title="Key Metrics")
        _add_section_title(ws, row=counts_by_level_title_row, title="Counts by Level")
        _add_section_title(ws, row=counts_by_service_title_row, title="Counts by Service")

        _style_header_row(ws, key_metrics_header_row)
        _style_header_row(ws, counts_by_level_header_row)
        _style_header_row(ws, counts_by_service_header_row)
        _style_header_row(ws, 2)  # metric/value
        _style_header_row(ws, 7)  # level/count
        _style_header_row(ws, 7 + per_level_len + 3)  # service/count

        # Optional: keep titles + first table header visible while scrolling
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

        _autofit_columns(ws)

    wb.save(xlsx_path)


# -------------------------
# Write the full report
# -------------------------


def write_excel_report(df: pd.DataFrame, output_path: Path) -> None:
    """
    Write the Excel report:
      - logs
      - summary (3 sections)
      - daily_summary

    Then apply styling as a second pass (reliable with openpyxl).
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df, per_level, per_service = build_summary_tables(df)
    daily_summary_df = build_daily_summary(df)
    logs_df = build_export_logs(df)

    # IMPORTANT:
    # We reserve row 1 for the "Key Metrics" title.
    # So summary_df starts at row 2 (startrow=1 in pandas, 0-indexed).
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        logs_df.to_excel(writer, sheet_name="logs", index=False)

        # summary sheet layout
        summary_df.to_excel(writer, sheet_name="summary", index=False, startrow=1)

        # Title at row 6 => per_level header should be row 7 => startrow=6
        per_level.to_excel(writer, sheet_name="summary", index=False, startrow=6)

        # Title at row (9 + per_level_len) => per_service header at next row => startrow=(9 + per_level_len)
        per_service_startrow = 9 + len(per_level)
        per_service.to_excel(
            writer, sheet_name="summary", index=False, startrow=per_service_startrow
        )

        daily_summary_df.to_excel(writer, sheet_name="daily_summary", index=False)

    # Style after saving
    style_workbook(output_path, per_level_len=len(per_level))

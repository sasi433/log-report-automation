from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from report_utils import (
    apply_filters,
    build_daily_summary,
    build_export_logs,
    build_summary_tables,
    load_logs,
)

EXIT_OK = 0
EXIT_ERROR = 1
EXIT_INPUT_MISSING = 2
EXIT_OUTPUT_ERROR = 3


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Log Report Automation - generate simple reports from CSV logs."
    )
    parser.add_argument("--input", default="sample_data/example.csv")
    parser.add_argument("--output", default="reports/report.xlsx")
    parser.add_argument("--service", default=None, help="Filter by service name (e.g., api, auth, db)")
    parser.add_argument("--level", default=None, help="Filter by level (e.g., INFO, ERROR)")
    return parser.parse_args()


def print_stats(df: pd.DataFrame) -> None:
    print("\n--- Basic stats ---")
    print(f"Total rows: {len(df)}")

    print("\nCount by level:")
    print(df["level"].value_counts(dropna=False).to_string())

    print("\nCount by service:")
    print(df["service"].value_counts(dropna=False).to_string())

    bad_ts = int(df["timestamp"].isna().sum())
    bad_ms = int(df["response_ms"].isna().sum())
    if bad_ts or bad_ms:
        print("\nData quality warnings:")
        if bad_ts:
            print(f"- Invalid timestamps: {bad_ts}")
        if bad_ms:
            print(f"- Invalid response_ms values: {bad_ms}")


def format_worksheet_columns(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def left_align_body(ws) -> None:
    left = Alignment(horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = left


def write_excel_report(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df, per_level, per_service = build_summary_tables(df)
    daily_summary_df = build_daily_summary(df)
    logs_df = build_export_logs(df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        logs_df.to_excel(writer, sheet_name="logs", index=False)

        summary_df.to_excel(writer, sheet_name="summary", index=False, startrow=0)
        per_level.to_excel(writer, sheet_name="summary", index=False, startrow=5)
        per_service.to_excel(writer, sheet_name="summary", index=False, startrow=5 + len(per_level) + 3)

        daily_summary_df.to_excel(writer, sheet_name="daily_summary", index=False)

        ws_logs = writer.sheets["logs"]
        ws_daily = writer.sheets["daily_summary"]
        ws_summary = writer.sheets["summary"]

        # logs formatting
        ws_logs.freeze_panes = "A2"
        ws_logs.auto_filter.ref = ws_logs.dimensions
        left_align_body(ws_logs)
        format_worksheet_columns(ws_logs)

        # daily formatting
        ws_daily.freeze_panes = "A2"
        ws_daily.auto_filter.ref = ws_daily.dimensions
        left_align_body(ws_daily)
        format_worksheet_columns(ws_daily)

        # summary (optional autosize)
        format_worksheet_columns(ws_summary)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    print("✅ Log Report Automation")
    print(f"Input : {input_path.resolve()}")
    print(f"Output: {output_path.resolve()}")

    if args.service or args.level:
        print("\n--- Active filters ---")
        if args.service:
            print(f"service = {args.service}")
        if args.level:
            print(f"level   = {args.level.upper()}")

    try:
        df = load_logs(input_path)
        df = apply_filters(df, args.service, args.level)
    except FileNotFoundError as exc:
        print(f"\n❌ {exc}")
        return EXIT_INPUT_MISSING
    except Exception as exc:
        print(f"\n❌ Error: {exc}")
        return EXIT_ERROR

    if df.empty:
        print("\n⚠️ No rows match the given filters. No report generated.")
        return EXIT_OK

    print_stats(df)

    try:
        write_excel_report(df, output_path)
    except Exception as exc:
        print(f"\n❌ Failed to write Excel report: {exc}")
        return EXIT_OUTPUT_ERROR

    print(f"\n✅ Excel report generated: {output_path.resolve()} (sheets: logs, summary, daily_summary)")
    return EXIT_OK


if __name__ == "__main__":
    raise SystemExit(main())

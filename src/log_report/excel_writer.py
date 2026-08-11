from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .analytics import (
    build_daily_summary,
    build_errors_by_service,
    build_export_logs,
    build_service_performance,
    build_summary_metrics,
)
from .validation import ValidationResult, build_issue_details, build_quality_summary

DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EDF4FA"
PALE_RED = "FCE8E6"
WHITE = "FFFFFF"
GRAY = "667085"
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FONT = Font(bold=True, color=WHITE)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def sanitize_excel_text(value: object) -> object:
    """Prevent openpyxl from turning untrusted text into an Excel formula.

    openpyxl marks strings beginning with ``=`` as formula cells. Strings
    beginning with ``+``, ``-``, or ``@`` remain ordinary string cells and are
    intentionally preserved.
    """
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def _sanitize_frame(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda column: column.map(sanitize_excel_text))


def _autofit_columns(ws, max_width: int = 60, scan_limit: int = 2000) -> None:
    for column in ws.columns:
        first_cell = column[0]
        if not hasattr(first_cell, "column_letter"):
            continue
        width = max(
            (len("" if cell.value is None else str(cell.value)) for cell in column[:scan_limit]),
            default=0,
        )
        ws.column_dimensions[first_cell.column_letter].width = min(width + 2, max_width)


def _style_header_row(ws, row: int, start_column: int = 1, end_column: int | None = None) -> None:
    end_column = end_column or ws.max_column
    for column in range(start_column, end_column + 1):
        cell = ws.cell(row=row, column=column)
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _style_title(ws, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=20, color=DARK_BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32


def _style_section_title(ws, row: int, title: str, start_column: int, end_column: int) -> None:
    ws.merge_cells(
        start_row=row,
        start_column=start_column,
        end_row=row,
        end_column=end_column,
    )
    cell = ws.cell(row=row, column=start_column, value=title)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = Font(bold=True, size=12, color=DARK_BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def _highlight_error_rows(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    if "level" not in headers or ws.max_row < 2:
        return
    level_column = ws.cell(row=1, column=headers.index("level") + 1).column_letter
    last_column = ws.cell(row=1, column=ws.max_column).column_letter
    rule = FormulaRule(
        formula=[f'${level_column}2="ERROR"'],
        fill=PatternFill("solid", fgColor=PALE_RED),
    )
    ws.conditional_formatting.add(f"A2:{last_column}{ws.max_row}", rule)


def _style_logs(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    _style_header_row(ws, 1)
    _highlight_error_rows(ws)
    _autofit_columns(ws)
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 16
    for cell in ws["F"][1:]:
        cell.number_format = '0" ms"'


def _add_errors_chart(ws, service_count: int) -> None:
    if service_count == 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Errors by Service"
    chart.height = 7
    chart.width = 12
    chart.gapWidth = 50
    chart.varyColors = False
    data = Reference(ws, min_col=9, min_row=17, max_row=17 + service_count)
    categories = Reference(ws, min_col=8, min_row=18, max_row=17 + service_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    chart.dLbls = DataLabelList(
        dLblPos="outEnd",
        showLegendKey=False,
        showVal=True,
        showCatName=True,
        showSerName=False,
        showPercent=False,
        showBubbleSize=False,
        separator=": ",
    )
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = MEDIUM_BLUE
        chart.series[0].graphicalProperties.line.solidFill = MEDIUM_BLUE
    ws.add_chart(chart, "D4")


def _style_summary(ws, service_count: int, error_service_count: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A4"
    _style_title(ws, "Log Report Summary", 9)
    ws["A2"] = "Operational reliability and latency overview"
    ws["A2"].font = Font(italic=True, color=GRAY)

    _style_section_title(ws, 3, "Key Metrics", 1, 2)
    _style_header_row(ws, 4, 1, 2)
    for row in range(5, 14):
        ws.cell(row=row, column=1).font = Font(bold=True, color=DARK_BLUE)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
    ws["B5"].number_format = "0"
    ws["B6"].number_format = "0"
    ws["B7"].number_format = '0.0"%"'
    for row in range(8, 12):
        ws.cell(row=row, column=2).number_format = '0.0" ms"'
    ws["B12"].number_format = "0"
    ws["B13"].number_format = '0" ms"'

    _style_section_title(ws, 16, "Service Performance", 1, 6)
    _style_header_row(ws, 17, 1, 6)
    _style_section_title(ws, 16, "Errors by Service", 8, 9)
    _style_header_row(ws, 17, 8, 9)
    for row in range(18, 18 + service_count):
        ws.cell(row=row, column=4).number_format = '0.0"%"'
        ws.cell(row=row, column=5).number_format = '0.0" ms"'
        ws.cell(row=row, column=6).number_format = '0.0" ms"'
        for column in range(1, 7):
            ws.cell(row=row, column=column).border = THIN_BORDER
    for row in range(18, 18 + error_service_count):
        for column in range(8, 10):
            ws.cell(row=row, column=column).border = THIN_BORDER

    _autofit_columns(ws)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 19
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    _add_errors_chart(ws, error_service_count)


def _add_daily_chart(ws) -> None:
    if ws.max_row < 2:
        return
    headers = [cell.value for cell in ws[1]]
    error_column = headers.index("error_count") + 1
    chart = LineChart()
    chart.style = 13
    chart.title = "Daily Error Trend"
    chart.y_axis.title = "Errors"
    chart.x_axis.title = "Date"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=error_column, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = MEDIUM_BLUE
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 6
    ws.add_chart(chart, "K2")


def _style_daily(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    _style_header_row(ws, 1, 1, 9)
    for cell in ws["D"][1:]:
        cell.number_format = '0.0"%"'
    for column in ("E", "F"):
        for cell in ws[column][1:]:
            cell.number_format = '0.0" ms"'
    for column in range(1, 10):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=column).border = THIN_BORDER
    _autofit_columns(ws)
    _add_daily_chart(ws)


def _style_quality(ws, detail_count: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A4"
    _style_title(ws, "Data Quality", 4)
    ws["A2"] = "Validation findings from the original CSV input"
    ws["A2"].font = Font(italic=True, color=GRAY)
    _style_section_title(ws, 3, "Validation Summary", 1, 2)
    _style_header_row(ws, 4, 1, 2)
    for row in range(5, 13):
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        if ws.cell(row=row, column=2).value:
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=PALE_RED)
            ws.cell(row=row, column=2).font = Font(bold=True, color="9C0006")
    _style_section_title(ws, 15, "Affected Rows", 1, 4)
    _style_header_row(ws, 16, 1, 4)
    for row in range(17, 17 + detail_count):
        for column in range(1, 5):
            ws.cell(row=row, column=column).border = THIN_BORDER
    _autofit_columns(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 38


def style_workbook(
    path: Path,
    service_count: int,
    error_service_count: int,
    detail_count: int,
) -> None:
    wb = load_workbook(path)
    wb.active = wb.sheetnames.index("summary")
    _style_summary(wb["summary"], service_count, error_service_count)
    _style_logs(wb["logs"])
    _style_daily(wb["daily_summary"])
    _style_quality(wb["data_quality"], detail_count)
    wb.save(path)


def write_excel_report(
    df: pd.DataFrame,
    output_path: Path,
    validation: ValidationResult | None = None,
    slow_threshold_ms: float = 500,
) -> None:
    """Write the stakeholder workbook and apply its presentation styling."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if validation is None:
        validation = ValidationResult(df, (), 0)

    metrics = _sanitize_frame(build_summary_metrics(df, slow_threshold_ms))
    service_performance = _sanitize_frame(build_service_performance(df))
    errors_by_service = _sanitize_frame(build_errors_by_service(df))
    logs = _sanitize_frame(build_export_logs(df))
    daily = _sanitize_frame(build_daily_summary(df))
    quality = _sanitize_frame(build_quality_summary(validation))
    issue_details = _sanitize_frame(build_issue_details(validation))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="summary", index=False, startrow=3)
        service_performance.to_excel(writer, sheet_name="summary", index=False, startrow=16)
        errors_by_service.to_excel(
            writer,
            sheet_name="summary",
            index=False,
            startrow=16,
            startcol=7,
        )
        logs.to_excel(writer, sheet_name="logs", index=False)
        daily.to_excel(writer, sheet_name="daily_summary", index=False)
        quality.to_excel(writer, sheet_name="data_quality", index=False, startrow=3)
        issue_details.to_excel(writer, sheet_name="data_quality", index=False, startrow=15)

    style_workbook(
        output_path,
        service_count=len(service_performance),
        error_service_count=len(errors_by_service),
        detail_count=len(issue_details),
    )

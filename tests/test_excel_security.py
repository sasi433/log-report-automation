from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from log_report.excel_writer import sanitize_excel_text, write_excel_report
from log_report.validation import normalize_and_validate


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("=2+2", "'=2+2"),
        ("+SUM(A1:A2)", "+SUM(A1:A2)"),
        ("-10+20", "-10+20"),
        ("@SUM(A1:A2)", "@SUM(A1:A2)"),
        ("ordinary message", "ordinary message"),
    ],
)
def test_sanitize_excel_text_only_changes_openpyxl_formula_strings(value, expected):
    assert sanitize_excel_text(value) == expected


def test_exported_formula_like_messages_are_safe_text(tmp_path):
    messages = ["=2+2", "+SUM(A1:A2)", "-10+20", "@SUM(A1:A2)"]
    source = pd.DataFrame(
        {
            "timestamp": ["2025-01-01T10:00:00Z"] * len(messages),
            "service": ['=WEBSERVICE("https://example.invalid")', "api", "api", "api"],
            "level": ["INFO"] * len(messages),
            "message": messages,
            "response_ms": [10, 20, 30, 40],
        }
    )
    result = normalize_and_validate(source)
    output = tmp_path / "safe.xlsx"

    write_excel_report(result.data, output, result)

    workbook = load_workbook(output, data_only=False)
    logs = workbook["logs"]
    exported_messages = [logs.cell(row=row, column=5) for row in range(2, 6)]
    assert [cell.value for cell in exported_messages] == [
        "'=2+2",
        "+SUM(A1:A2)",
        "-10+20",
        "@SUM(A1:A2)",
    ]
    assert all(cell.data_type == "s" for cell in exported_messages)
    assert logs["C2"].value.startswith("'=")
    assert logs["C2"].data_type == "s"

    formula_cells = [
        cell.coordinate
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert formula_cells == []

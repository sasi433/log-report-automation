from __future__ import annotations

import sys

from openpyxl import load_workbook

from log_report.cli import EXIT_ERROR, EXIT_INPUT_MISSING, EXIT_OK, main


def test_cli_missing_input_returns_distinct_exit_code(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(
        sys,
        "argv",
        ["log-report", "--input", str(tmp_path / "missing.csv")],
    )

    assert main() == EXIT_INPUT_MISSING
    assert "Input file not found" in capsys.readouterr().out


def test_cli_empty_filter_does_not_create_report(tmp_path, monkeypatch, capsys):
    output_path = tmp_path / "report.xlsx"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "log-report",
            "--input",
            "sample_data/example.csv",
            "--output",
            str(output_path),
            "--service",
            "missing-service",
        ],
    )

    assert main() == EXIT_OK
    assert not output_path.exists()
    assert "No usable rows match" in capsys.readouterr().out


def write_invalid_csv(path):
    path.write_text(
        "timestamp,service,level,message,response_ms\n"
        "2025-01-01T10:00:00Z,api,INFO,ok,25\n"
        "bad-date,api,ERROR,bad timestamp,50\n"
        "2025-01-01T11:00:00Z,,WARN,no service,75\n",
        encoding="utf-8",
    )


def test_cli_strict_validation_fails_without_report(tmp_path, capsys):
    input_path = tmp_path / "invalid.csv"
    output_path = tmp_path / "report.xlsx"
    write_invalid_csv(input_path)

    exit_code = main(["--input", str(input_path), "--output", str(output_path)])

    assert exit_code == EXIT_ERROR
    assert not output_path.exists()
    assert "Validation failed: 2 invalid rows" in capsys.readouterr().out


def test_cli_lenient_validation_writes_quality_sheet(tmp_path, capsys):
    input_path = tmp_path / "invalid.csv"
    output_path = tmp_path / "report.xlsx"
    write_invalid_csv(input_path)

    exit_code = main(
        [
            "--input",
            str(input_path),
            "--output",
            str(output_path),
            "--validation",
            "lenient",
        ]
    )

    assert exit_code == EXIT_OK
    assert output_path.exists()
    output = capsys.readouterr().out
    assert "Rejected rows: 2" in output
    assert "Sheets: summary, logs, daily_summary, data_quality" in output
    assert "data_quality" in load_workbook(output_path).sheetnames


def test_cli_applies_custom_slow_threshold(tmp_path):
    output_path = tmp_path / "report.xlsx"

    exit_code = main(
        [
            "--input",
            "sample_data/example.csv",
            "--output",
            str(output_path),
            "--slow-threshold-ms",
            "123",
        ]
    )

    assert exit_code == EXIT_OK
    assert load_workbook(output_path)["summary"]["B13"].value == 123

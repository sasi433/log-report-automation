from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from log_report.excel_writer import write_excel_report
from log_report.validation import load_logs


def test_write_excel_report_creates_expected_sheets(tmp_path: Path) -> None:
    input_csv = Path("sample_data/example.csv")
    output_xlsx = tmp_path / "report.xlsx"

    result = load_logs(input_csv)
    write_excel_report(result.data, output_xlsx, result)

    assert output_xlsx.exists()

    wb = load_workbook(output_xlsx)
    assert wb.sheetnames == ["summary", "logs", "daily_summary", "data_quality"]

    # logs sheet: header sanity
    ws_logs = wb["logs"]
    headers = [cell.value for cell in ws_logs[1]]
    assert headers == ["date", "time", "service", "level", "message", "response_ms"]

    # logs sheet: should contain data rows
    assert ws_logs.max_row >= 2

    # summary: key metrics table exists
    ws_summary = wb["summary"]
    assert ws_summary["A1"].value == "Log Report Summary"
    # these should exist somewhere near the top (robust checks)
    values = [ws_summary.cell(row=r, column=1).value for r in range(1, 15)]
    assert "metric" in values or "Key Metrics" in values

    # daily_summary: has date column
    ws_daily = wb["daily_summary"]
    daily_headers = [cell.value for cell in ws_daily[1]]
    assert "date" in daily_headers

    ws_quality = wb["data_quality"]
    assert ws_quality["A1"].value == "Data Quality"
    assert wb.active.title == "summary"
